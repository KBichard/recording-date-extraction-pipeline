from pathlib import Path

import pandas as pd

from datetime import datetime

 
# ============================================================
# PATH CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

LOOKUP_PATH    = BASE_DIR / "data" / "county_list.csv"
RECORDING_PATH = BASE_DIR / "Output" / "recording_dates_predictions.xlsx"
PROCESSED_PATH = BASE_DIR / "data" / "processed_dates.xlsx"
FINAL_PATH     = BASE_DIR / "Output" / "county_dashboard_master.xlsx"

 

# ============================================================

# LOAD DATA  ✅ (CORRECT)

# ============================================================

 

lookup_df    = pd.read_csv(LOOKUP_PATH, dtype=str)

recording_df = pd.read_excel(RECORDING_PATH, dtype=str)

processed_df = pd.read_excel(PROCESSED_PATH, dtype=str) 

 

# ============================================================

# NORMALIZE LOOKUP DATA (ALL CAPS)

# ============================================================

 

lookup_df["SUBCODE"] = lookup_df["SUBCODE"].str.strip().str.upper()

lookup_df["FIPS"] = lookup_df["FIPS"].str.replace(r"\D", "", regex=True)

lookup_df["FREQUENCY"] = pd.to_numeric(lookup_df["FREQUENCY"], errors="coerce")

lookup_df["PROCESSING TYPE"] = lookup_df["PROCESSING TYPE"].str.strip()

 

# ============================================================

# NORMALIZE PROCESSED DATA (AUTHORITATIVE OCR SOURCE)

# ============================================================

 

processed_df["SUBCODE"] = processed_df["SUBCODE"].str.strip().str.upper()

processed_df["Processed Date"] = pd.to_datetime(

    processed_df["Processed Date"], errors="coerce"

)

 

# ============================================================

# NORMALIZE RECORDING DATA

# ============================================================

 

recording_df["FIPS"] = recording_df["FIPS"].str.replace(r"\D", "", regex=True)

 

# ============================================================

# ATTACH COUNTY METADATA TO PROCESSED DATA

# ============================================================

 

processed_df = processed_df.merge(

    lookup_df[

        [

            "FIPS",

            "SUBCODE",

            "STATE",

            "COUNTY",

            "FREQUENCY",

            "PROCESSING TYPE",

        ]

    ],

    on="SUBCODE",

    how="left"

)

 

# ============================================================

# MERGE RECORDING DATE PREDICTIONS

# ============================================================

 

final_df = processed_df.merge(

    recording_df[

        [

            "FIPS",

            "DATE_STRING",

            "RECORDING_DATE_SCORE",

            "DECISION_STATUS",

        ]

    ],

    on="FIPS",

    how="left"

)

 

 

# ============================================================

# SLA / FREQUENCY CALCULATIONS

# ============================================================

 

today = pd.Timestamp(datetime.today().date())

 

final_df["Days Since Processed"] = (

    today - final_df["Processed Date"]

).dt.days.clip(lower=0)

 

final_df["Overdue By (Days)"] = (

    final_df["Days Since Processed"] - final_df["FREQUENCY"]

).clip(lower=0)

 

final_df["Update Status"] = final_df["Overdue By (Days)"].apply(

    lambda x: "OVERDUE" if x > 0 else "OK"

)

 

 

 

# ============================================================

# CLEANUP, RENAME, FINAL ORDER

# ============================================================

 

final_df = final_df.rename(

    columns={

        "DATE_STRING": "Recording Date",

        "RECORDING_DATE_SCORE": "Recording Confidence",

        "DECISION_STATUS": "Recording Decision",

        "FREQUENCY": "Expected Frequency (Days)",

    }

)

 

final_df = final_df[

    [

        "FIPS",

        "STATE",

        "COUNTY",

        "SUBCODE",

        "PROCESSING TYPE",

        "Latest Image",

        "Processed Date",

        "Days Since Processed",

        "Expected Frequency (Days)",

        "Overdue By (Days)",

        "Update Status",

        "Recording Date",

        "Recording Confidence",

        "Recording Decision",

    ]

]

 

 

from openpyxl import load_workbook

from openpyxl.styles import PatternFill

from openpyxl.formatting.rule import CellIsRule, FormulaRule

 

# ============================================================

# SAVE OUTPUT (WITH FORMATTING)

# ============================================================

 

final_df.to_excel(FINAL_PATH, index=False)

 

wb = load_workbook(FINAL_PATH)

ws = wb.active

 

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

 

headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

 

overdue_col = headers["Overdue By (Days)"]

status_col = headers["Update Status"]

 

ws.conditional_formatting.add(

    f"{chr(64+overdue_col)}2:{chr(64+overdue_col)}{ws.max_row}",

    CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill)

)

 

ws.conditional_formatting.add(

    f"{chr(64+status_col)}2:{chr(64+status_col)}{ws.max_row}",

    FormulaRule(formula=[f'${chr(64+status_col)}2="OVERDUE"'], fill=red_fill)

)

 

wb.save(FINAL_PATH)

 

print("Master County Dashboard updated successfully:")

print(FINAL_PATH)
